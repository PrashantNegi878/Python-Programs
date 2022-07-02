import openpyxl as xl

# import random
# for i in range (5):
#     print(random.randint(10,50))
wb=xl.load_workbook("transactions.xlsx")
sheet=wb['Sheet1']

# for col in range (1,4):
for row in range (2,5):
    cell=sheet.cell(row,3)
    cell4=sheet.cell(row,4)
    correctprice=sheet.cell(row,3).value*10
    cell4.value=correctprice
wb.save('transnew.xlsx')
